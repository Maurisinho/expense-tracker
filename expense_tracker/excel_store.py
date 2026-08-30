"""Almacenamiento de las transacciones en un libro Excel (xlsx)."""

import os
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Transaction

HOJA = "Transacciones"
COLUMNAS = ["id", "tipo", "descripcion", "monto", "categoria", "fecha"]


def _ruta_one_drive() -> Path:
    for variable in ("OneDriveConsumer", "OneDrive"):
        valor = os.environ.get(variable)
        if valor and os.path.isdir(valor):
            return Path(valor)
    return Path.home() / "OneDrive"


def default_excel_file() -> Path:
    """Devuelve la ruta del Excel de datos en tu carpeta de OneDrive."""
    return _ruta_one_drive() / "ExpenseTracker" / "transacciones.xlsx"


class ExcelStorage:
    """Lee y escribe las transacciones en un archivo Excel."""

    def __init__(self, path) -> None:
        self.path = Path(path)

    def load(self) -> List[Transaction]:
        """Carga las transacciones desde la hoja del Excel."""
        if not self.path.exists():
            return []
        workbook = load_workbook(self.path, data_only=True)
        try:
            sheet = workbook.active
            transacciones: List[Transaction] = []
            for fila in sheet.iter_rows(min_row=2, values_only=True):
                if fila[0] is None:
                    continue
                raw = dict(zip(COLUMNAS, fila))
                if not isinstance(raw["fecha"], str):
                    raw["fecha"] = raw["fecha"].isoformat()
                transacciones.append(Transaction.from_dict(raw))
        finally:
            workbook.close()
        return transacciones

    def save(self, transactions: List[Transaction]) -> None:
        """Escribe todas las transacciones en el Excel (primera columna: cabecera)."""
        self.path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        try:
            sheet = workbook.active
            sheet.title = HOJA
            relleno = PatternFill("solid", fgColor="E8EAF6")
            for indice, titulo in enumerate(COLUMNAS, start=1):
                celda = sheet.cell(row=1, column=indice, value=titulo)
                celda.font = Font(bold=True)
                celda.fill = relleno
                sheet.column_dimensions[get_column_letter(indice)].width = max(
                    14, len(titulo) + 2
                )
            for tx in transactions:
                valores = [tx.id, tx.tipo, tx.descripcion, tx.monto, tx.categoria,
                           tx.fecha.isoformat()]
                sheet.append(valores)
                fila_actual = sheet.max_row
                sheet.cell(row=fila_actual, column=4).number_format = "#,##0.00"
            sheet.freeze_panes = "A2"
            workbook.save(self.path)
        finally:
            workbook.close()